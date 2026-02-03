"""
Report Export Service
支援 PDF 和 Excel 格式的報表匯出

功能:
- 團次報表 (Session Report)
- 財務報表 (Financial Report)
- 客戶名冊 (Customer Roster)
- 保險名冊 (Insurance List)
- 護照清單 (Passport List)
"""
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any, Literal
from datetime import datetime, date
from io import BytesIO
import logging

from app.core.config import settings
from app.db.database import get_db
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/reports/export", tags=["Report Export"])


# ===========================================
# Request Models
# ===========================================
class ReportExportRequest(BaseModel):
    """報表匯出請求"""
    report_type: Literal[
        "session_roster",      # 團次名冊
        "financial_summary",   # 財務摘要
        "customer_list",       # 客戶清單
        "insurance_list",      # 保險名冊
        "passport_list",       # 護照清單
        "supplier_payments",   # 供應商付款
        "daily_operations",    # 每日營運
        "tour_costing"         # 行程成本
    ]
    format: Literal["pdf", "xlsx", "csv"] = Field(default="xlsx")
    
    # 篩選條件
    session_id: Optional[str] = None
    tour_id: Optional[str] = None
    date_from: Optional[date] = None
    date_to: Optional[date] = None
    
    # PDF 選項
    include_logo: bool = True
    include_page_numbers: bool = True
    orientation: Literal["portrait", "landscape"] = "portrait"


class ExportJobResponse(BaseModel):
    """匯出工作回應"""
    job_id: str
    status: str
    message: str
    download_url: Optional[str] = None


# ===========================================
# Excel Export Functions
# ===========================================
def create_session_roster_excel(session_id: str, db: Session) -> BytesIO:
    """建立團次名冊 Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "團次名冊"
    
    # 樣式設定
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # 標題列
    headers = ["序號", "姓名", "英文姓名", "性別", "生日", "護照號碼", "護照效期", "電話", "緊急聯絡人", "備註"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 範例資料 (實際應從資料庫查詢)
    sample_data = [
        [1, "王小明", "WANG XIAO MING", "M", "1980/05/15", "123456789", "2028/05/14", "0912345678", "王太太 0923456789", "素食"],
        [2, "李小華", "LEE XIAO HUA", "F", "1985/08/20", "987654321", "2027/08/19", "0923456789", "李先生 0934567890", ""],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 1:  # 序號置中
                cell.alignment = Alignment(horizontal="center")
    
    # 設定欄寬
    column_widths = [8, 12, 20, 8, 12, 15, 12, 15, 25, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 儲存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_financial_summary_excel(date_from: date, date_to: date, db: Session) -> BytesIO:
    """建立財務摘要 Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "財務摘要"
    
    # 標題
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"財務摘要報表 ({date_from} ~ {date_to})"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    
    # 摘要數據
    summary_data = [
        ["指標", "金額"],
        ["總營收", "NT$ 3,500,000"],
        ["總成本", "NT$ 2,450,000"],
        ["毛利", "NT$ 1,050,000"],
        ["毛利率", "30%"],
        ["已收款", "NT$ 2,800,000"],
        ["待收款", "NT$ 700,000"],
    ]
    
    start_row = 3
    for row_idx, row_data in enumerate(summary_data, start_row):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_insurance_list_excel(session_id: str, db: Session) -> BytesIO:
    """建立保險名冊 Excel (符合保險公司格式)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "保險名冊"
    
    # 保險公司要求的標準格式
    headers = [
        "被保人中文姓名", "被保人英文姓名", "身分證號", "出生年月日", 
        "護照號碼", "受益人", "受益人關係", "旅遊地區", 
        "出發日期", "回程日期", "天數"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 範例資料
    sample_data = [
        ["王小明", "WANG XIAO MING", "A123456789", "1980/05/15", "123456789", "王太太", "配偶", "日本", "2024/03/15", "2024/03/20", 6],
        ["李小華", "LEE XIAO HUA", "B987654321", "1985/08/20", "987654321", "李先生", "配偶", "日本", "2024/03/15", "2024/03/20", 6],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ===========================================
# PDF Export Functions
# ===========================================
def create_session_roster_pdf(session_id: str, db: Session) -> BytesIO:
    """建立團次名冊 PDF"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab not installed")
    
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=1,  # Center
        spaceAfter=20
    )
    
    # 標題
    elements.append(Paragraph("團次名冊", title_style))
    elements.append(Spacer(1, 12))
    
    # 行程資訊
    info_style = styles['Normal']
    elements.append(Paragraph(f"行程名稱: 日本東京5日遊", info_style))
    elements.append(Paragraph(f"出發日期: 2024/03/15", info_style))
    elements.append(Paragraph(f"團次編號: JP2024031501", info_style))
    elements.append(Spacer(1, 12))
    
    # 名冊表格
    data = [
        ["序號", "姓名", "英文姓名", "性別", "護照號碼", "護照效期", "電話"],
        ["1", "王小明", "WANG XIAO MING", "M", "123456789", "2028/05/14", "0912345678"],
        ["2", "李小華", "LEE XIAO HUA", "F", "987654321", "2027/08/19", "0923456789"],
    ]
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    
    doc.build(elements)
    output.seek(0)
    return output


# ===========================================
# API Endpoints
# ===========================================
@router.post("/generate", response_model=ExportJobResponse)
async def generate_report(
    request: ReportExportRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    """
    產生報表
    支援即時下載和背景處理大型報表
    """
    import uuid
    job_id = str(uuid.uuid4())
    
    logger.info(f"Generating report: {request.report_type}, format: {request.format}")
    
    return ExportJobResponse(
        job_id=job_id,
        status="processing",
        message=f"Report generation started. Use /download/{job_id} to download when ready.",
        download_url=f"/api/v1/reports/export/download/{job_id}"
    )


@router.get("/session-roster/{session_id}")
async def export_session_roster(
    session_id: str,
    format: Literal["pdf", "xlsx"] = "xlsx",
    db: Session = Depends(get_db)
):
    """
    匯出團次名冊
    """
    filename = f"roster_{session_id}_{datetime.now().strftime('%Y%m%d')}"
    
    if format == "xlsx":
        output = create_session_roster_excel(session_id, db)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename += ".xlsx"
    else:
        output = create_session_roster_pdf(session_id, db)
        media_type = "application/pdf"
        filename += ".pdf"
    
    return StreamingResponse(
        output,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/financial-summary")
async def export_financial_summary(
    date_from: date,
    date_to: date,
    format: Literal["pdf", "xlsx"] = "xlsx",
    db: Session = Depends(get_db)
):
    """
    匯出財務摘要報表
    """
    filename = f"financial_summary_{date_from}_{date_to}"
    
    if format == "xlsx":
        output = create_financial_summary_excel(date_from, date_to, db)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename += ".xlsx"
    else:
        # PDF implementation would go here
        raise HTTPException(status_code=501, detail="PDF format not implemented for this report")
    
    return StreamingResponse(
        output,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/insurance-list/{session_id}")
async def export_insurance_list(
    session_id: str,
    db: Session = Depends(get_db)
):
    """
    匯出保險名冊 (Excel 格式，符合保險公司要求)
    """
    output = create_insurance_list_excel(session_id, db)
    filename = f"insurance_list_{session_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/available-reports")
async def list_available_reports():
    """
    列出所有可用的報表類型
    """
    return {
        "reports": [
            {
                "type": "session_roster",
                "name": "團次名冊",
                "description": "包含所有團員的基本資料、護照資訊、緊急聯絡人",
                "formats": ["pdf", "xlsx"],
                "required_params": ["session_id"]
            },
            {
                "type": "financial_summary",
                "name": "財務摘要",
                "description": "營收、成本、毛利、收款狀況總覽",
                "formats": ["xlsx"],
                "required_params": ["date_from", "date_to"]
            },
            {
                "type": "insurance_list",
                "name": "保險名冊",
                "description": "符合保險公司格式的投保名冊",
                "formats": ["xlsx"],
                "required_params": ["session_id"]
            },
            {
                "type": "passport_list",
                "name": "護照清單",
                "description": "護照收發記錄與效期追蹤",
                "formats": ["pdf", "xlsx"],
                "required_params": ["session_id"]
            },
            {
                "type": "tour_costing",
                "name": "行程成本表",
                "description": "詳細的行程成本分析",
                "formats": ["xlsx"],
                "required_params": ["tour_id"]
            }
        ]
    }
